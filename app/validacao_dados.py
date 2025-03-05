import pandas as pd
import os
from pathlib import Path
import logging
from typing import Dict, List, Any, Tuple
from datetime import datetime
import numpy as np
import warnings
from pandas.errors import ParserError
import xlrd
from openpyxl import load_workbook
import json
from openpyxl.utils.exceptions import InvalidFileException

class ValidadorAvancado:
    def __init__(self):
        # Configuração de logging
        self.setup_logging()
        
        # Configuração de diretórios
        self.setup_directories()
        
        # Mapeamento de colunas com múltiplas estratégias
        self.config = self.carregar_configuracao()
        
        # Cache para otimização
        self.cache_dados = {}

    def setup_logging(self):
        """Configuração avançada de logging"""
        log_dir = Path('logs')
        log_dir.mkdir(exist_ok=True)
        
        self.logger = logging.getLogger('ValidadorAvancado')
        self.logger.setLevel(logging.DEBUG)
        
        # Handler para arquivo
        fh = logging.FileHandler(log_dir / f'validacao_{datetime.now().strftime("%Y%m%d_%H%M%S")}.log')
        fh.setLevel(logging.DEBUG)
        
        # Handler para console
        ch = logging.StreamHandler()
        ch.setLevel(logging.INFO)
        
        # Formatação
        formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
        fh.setFormatter(formatter)
        ch.setFormatter(formatter)
        
        self.logger.addHandler(fh)
        self.logger.addHandler(ch)

    def setup_directories(self):
        """Criação e verificação de diretórios necessários"""
        dirs = ['logs', 'config', 'cache', 'output']
        for d in dirs:
            Path(d).mkdir(exist_ok=True)

    def carregar_configuracao(self) -> Dict:
        """Carrega configuração com múltiplas estratégias de leitura"""
        config_path = Path('config/colunas_config.json')
        if not config_path.exists():
            config = {
                'colunas_essenciais': {
                    'DATA': {
                        'alternativas': ['DATA', 'DT', 'DATA_', 'DATA CADASTRO'],
                        'tipo': 'data',
                        'formato_data': '%Y-%m-%d',
                        'estrategias': ['direto', 'normalizado', 'fuzzy']
                    },
                    'RESOLUCAO': {
                        'alternativas': ['RESOLUÇÃO', 'RESOLUCAO', 'RESOL', 'DT_RESOLUCAO'],
                        'tipo': 'data',
                        'formato_data': '%Y-%m-%d',
                        'estrategias': ['direto', 'normalizado', 'fuzzy']
                    },
                    'CONTRATO': {
                        'alternativas': ['CONTRATO', 'NUM_CONTRATO', 'NÚMERO CONTRATO'],
                        'tipo': 'texto',
                        'estrategias': ['direto', 'normalizado']
                    },
                    'NEGOCIACAO': {
                        'alternativas': ['NEGOCIAÇÃO', 'NEGOCIACAO', 'DT_NEGOCIACAO'],
                        'tipo': 'data',
                        'formato_data': '%Y-%m-%d',
                        'estrategias': ['direto', 'normalizado', 'fuzzy']
                    },
                    'SITUACAO': {
                        'alternativas': ['SITUAÇÃO', 'SITUACAO', 'STATUS'],
                        'tipo': 'texto',
                        'valores_validos': [
                            'PENDENTE', 'QUITADO', 'APREENDIDO', 'PRIORIDADE',
                            'VERIFICADO', 'APROVADO', 'ANÁLISE'
                        ],
                        'estrategias': ['direto', 'normalizado']
                    }
                },
                'estrategias_leitura': ['pandas', 'openpyxl', 'xlrd'],
                'tratamento_erros': {
                    'datas_invalidas': 'corrigir',
                    'valores_vazios': 'preencher',
                    'formato_invalido': 'converter'
                }
            }
            with open(config_path, 'w', encoding='utf-8') as f:
                json.dump(config, f, indent=4, ensure_ascii=False)
        else:
            with open(config_path, 'r', encoding='utf-8') as f:
                config = json.load(f)
        return config

    def ler_planilha(self, arquivo: Path, aba: str) -> Tuple[pd.DataFrame, Dict[str, Any]]:
        """Lê planilha com múltiplas estratégias e tratamento de erros"""
        erros = []
        dados = None
        
        for estrategia in self.config['estrategias_leitura']:
            try:
                if estrategia == 'pandas':
                    dados = pd.read_excel(arquivo, sheet_name=aba)
                elif estrategia == 'openpyxl':
                    wb = load_workbook(arquivo, data_only=True)
                    ws = wb[aba]
                    dados = pd.DataFrame(ws.values)
                elif estrategia == 'xlrd':
                    wb = xlrd.open_workbook(arquivo)
                    ws = wb.sheet_by_name(aba)
                    dados = pd.DataFrame([ws.row_values(i) for i in range(ws.nrows)])
                
                if dados is not None and not dados.empty:
                    dados = self.limpar_dados(dados)
                    break
                    
            except Exception as e:
                erros.append(f"Erro na estratégia {estrategia}: {str(e)}")
                continue
        
        if dados is None:
            raise Exception(f"Todas as estratégias falharam: {'; '.join(erros)}")
            
        return dados, {'erros': erros}

    def limpar_dados(self, df: pd.DataFrame) -> pd.DataFrame:
        """Limpa e padroniza os dados"""
        # Remover linhas totalmente vazias
        df = df.dropna(how='all')
        
        # Identificar cabeçalho real
        for i in range(min(5, len(df))):
            potential_header = df.iloc[i]
            if self.verificar_cabecalho(potential_header):
                df.columns = potential_header
                df = df.iloc[i+1:].reset_index(drop=True)
                break
        
        # Limpar nomes das colunas
        df.columns = [str(col).strip() for col in df.columns]
        
        return df

    def verificar_cabecalho(self, row: pd.Series) -> bool:
        """Verifica se uma linha é provavelmente o cabeçalho"""
        colunas_importantes = set()
        for col in self.config['colunas_essenciais'].keys():
            alternativas = self.config['colunas_essenciais'][col]['alternativas']
            colunas_importantes.update(alt.upper() for alt in alternativas)
        
        row_values = set(str(val).upper().strip() for val in row if pd.notna(val))
        return len(row_values.intersection(colunas_importantes)) >= 2

    def processar_arquivo(self, arquivo: Path) -> Dict[str, Any]:
        """Processa um arquivo Excel completo"""
        self.logger.info(f"Processando arquivo: {arquivo}")
        resultados = {
            'arquivo': str(arquivo),
            'abas': {},
            'erros': [],
            'status': 'sucesso'
        }
        
        try:
            xls = pd.ExcelFile(arquivo)
            for aba in xls.sheet_names:
                if aba.upper() in ['TESTE', 'RESUMO', 'RELATÓRIO GERAL']:
                    continue
                    
                try:
                    self.logger.info(f"Processando aba: {aba}")
                    df, info = self.ler_planilha(arquivo, aba)
                    dados_processados = self.processar_aba(df, aba)
                    resultados['abas'][aba] = dados_processados
                    
                except Exception as e:
                    msg = f"Erro ao processar aba {aba}: {str(e)}"
                    self.logger.error(msg)
                    resultados['erros'].append(msg)
                    
        except Exception as e:
            msg = f"Erro ao processar arquivo: {str(e)}"
            self.logger.error(msg)
            resultados['erros'].append(msg)
            resultados['status'] = 'erro'
            
        return resultados

    def processar_aba(self, df: pd.DataFrame, nome_aba: str) -> Dict[str, Any]:
        """Processa uma aba individual com tratamento avançado de erros"""
        resultados = {
            'metricas': {
                'total_linhas': len(df),
                'total_colunas': len(df.columns),
                'colunas_encontradas': [],
                'problemas': []
            },
            'dados_validos': True
        }
        
        # Processar cada coluna essencial
        for col_nome, config in self.config['colunas_essenciais'].items():
            coluna_encontrada = self.encontrar_coluna(df, config['alternativas'])
            
            if coluna_encontrada:
                resultados['metricas']['colunas_encontradas'].append(col_nome)
                
                # Validar e corrigir dados
                if config['tipo'] == 'data':
                    df[coluna_encontrada] = self.corrigir_datas(df[coluna_encontrada])
                elif col_nome == 'SITUACAO':
                    df[coluna_encontrada] = self.validar_situacao(df[coluna_encontrada])
                    
            else:
                resultados['metricas']['problemas'].append(f"Coluna não encontrada: {col_nome}")
                resultados['dados_validos'] = False
        
        return resultados

    def encontrar_coluna(self, df: pd.DataFrame, alternativas: List[str]) -> str:
        """Encontra a coluna correta usando diferentes estratégias"""
        for col in df.columns:
            col_norm = self.normalizar_texto(str(col))
            for alt in alternativas:
                alt_norm = self.normalizar_texto(alt)
                if col_norm == alt_norm:
                    return col
        return None

    def normalizar_texto(self, texto: str) -> str:
        """Normaliza texto para comparação"""
        import unicodedata
        texto = unicodedata.normalize('NFKD', texto)
        texto = texto.encode('ASCII', 'ignore').decode('ASCII')
        return texto.upper().strip().replace(' ', '').replace('_', '')

    def corrigir_datas(self, serie: pd.Series) -> pd.Series:
        """Corrige problemas comuns em datas"""
        try:
            return pd.to_datetime(serie, errors='coerce', format='mixed')
        except:
            return pd.Series([None] * len(serie))

    def validar_situacao(self, serie: pd.Series) -> pd.Series:
        """Valida e corrige valores da coluna situação"""
        valores_validos = set(self.config['colunas_essenciais']['SITUACAO']['valores_validos'])
        serie = serie.fillna('PENDENTE')
        serie = serie.astype(str).str.upper().str.strip()
        
        # Corrigir valores inválidos
        mascara_invalidos = ~serie.isin(valores_validos)
        if mascara_invalidos.any():
            self.logger.warning(f"Valores inválidos encontrados em SITUACAO: {serie[mascara_invalidos].unique()}")
            serie[mascara_invalidos] = 'PENDENTE'
            
        return serie

class ExcelLeitor:
    def __init__(self):
        self.setup_logging()
        self.colunas_importantes = [
            'DATA', 'RESOLUÇÃO', 'CONTRATO', 'NEGOCIAÇÃO', 'SITUAÇÃO'
        ]
        
    def setup_logging(self):
        """Configura logging"""
        log_dir = Path('logs')
        log_dir.mkdir(exist_ok=True)
        
        logging.basicConfig(
            filename=log_dir / f'excel_leitor_{datetime.now().strftime("%Y%m%d")}.log',
            level=logging.INFO,
            format='%(asctime)s - %(levelname)s - %(message)s'
        )
        self.logger = logging.getLogger(__name__)

    def ler_arquivo(self, caminho_arquivo: str) -> Dict[str, Any]:
        """Lê arquivo Excel usando openpyxl"""
        try:
            self.logger.info(f"Iniciando leitura do arquivo: {caminho_arquivo}")
            resultado = {
                'status': 'sucesso',
                'arquivo': caminho_arquivo,
                'abas': {},
                'erros': []
            }

            # Carregar workbook com data_only=True para valores calculados
            wb = load_workbook(filename=caminho_arquivo, data_only=True, read_only=True)
            
            for nome_aba in wb.sheetnames:
                # Pular abas de resumo
                if nome_aba.upper() in ['RELATÓRIO GERAL', 'TESTE', 'RESUMO']:
                    continue
                    
                try:
                    dados_aba = self.processar_aba(wb[nome_aba])
                    if dados_aba:
                        resultado['abas'][nome_aba] = dados_aba
                except Exception as e:
                    msg = f"Erro ao processar aba {nome_aba}: {str(e)}"
                    self.logger.error(msg)
                    resultado['erros'].append(msg)

            wb.close()
            return resultado

        except InvalidFileException:
            msg = f"Arquivo inválido ou corrompido: {caminho_arquivo}"
            self.logger.error(msg)
            return {'status': 'erro', 'mensagem': msg}
        except Exception as e:
            msg = f"Erro ao ler arquivo: {str(e)}"
            self.logger.error(msg)
            return {'status': 'erro', 'mensagem': msg}

    def processar_aba(self, ws) -> Optional[Dict[str, Any]]:
        """Processa uma aba do Excel"""
        try:
            # Ler todas as linhas da aba
            rows = list(ws.rows)
            if not rows:
                return None

            # Encontrar cabeçalho
            header_row = self.encontrar_cabecalho(rows)
            if header_row is None:
                return None

            # Extrair cabeçalho e dados
            headers = [str(cell.value).strip() if cell.value else '' for cell in rows[header_row]]
            
            # Criar dicionário de índices das colunas importantes
            indices_colunas = self.mapear_colunas(headers)
            
            # Processar dados
            dados = []
            for row in rows[header_row + 1:]:
                linha = {}
                for col_nome, idx in indices_colunas.items():
                    if idx is not None:
                        valor = row[idx].value
                        linha[col_nome] = self.limpar_valor(valor, col_nome)
                if any(linha.values()):  # Adicionar apenas linhas não vazias
                    dados.append(linha)

            return {
                'total_linhas': len(dados),
                'colunas_encontradas': list(indices_colunas.keys()),
                'dados': dados
            }

        except Exception as e:
            self.logger.error(f"Erro ao processar aba {ws.title}: {str(e)}")
            return None

    def encontrar_cabecalho(self, rows: List) -> Optional[int]:
        """Encontra a linha do cabeçalho"""
        for idx, row in enumerate(rows[:10]):  # Verificar primeiras 10 linhas
            valores = [str(cell.value).strip().upper() if cell.value else '' for cell in row]
            if any(col in valores for col in ['DATA', 'RESOLUÇÃO', 'CONTRATO']):
                return idx
        return None

    def mapear_colunas(self, headers: List[str]) -> Dict[str, Optional[int]]:
        """Mapeia as colunas importantes para seus índices"""
        mapeamento = {
            'DATA': None,
            'RESOLUCAO': None,
            'CONTRATO': None,
            'NEGOCIACAO': None,
            'SITUACAO': None
        }
        
        alternativas = {
            'DATA': ['DATA', 'DT', 'DATA_'],
            'RESOLUCAO': ['RESOLUÇÃO', 'RESOLUCAO', 'RESOL'],
            'CONTRATO': ['CONTRATO', 'NUM_CONTRATO', 'NÚMERO CONTRATO'],
            'NEGOCIACAO': ['NEGOCIAÇÃO', 'NEGOCIACAO', 'DT_NEGOCIACAO'],
            'SITUACAO': ['SITUAÇÃO', 'SITUACAO', 'STATUS']
        }

        for idx, header in enumerate(headers):
            header_norm = self.normalizar_texto(header)
            for col_nome, alts in alternativas.items():
                if any(self.normalizar_texto(alt) == header_norm for alt in alts):
                    mapeamento[col_nome] = idx
                    break

        return mapeamento

    def normalizar_texto(self, texto: str) -> str:
        """Normaliza texto para comparação"""
        if not isinstance(texto, str):
            texto = str(texto)
        import unicodedata
        texto = unicodedata.normalize('NFKD', texto)
        texto = texto.encode('ASCII', 'ignore').decode('ASCII')
        return texto.upper().strip().replace(' ', '').replace('_', '')

    def limpar_valor(self, valor: Any, tipo_coluna: str) -> Any:
        """Limpa e formata o valor baseado no tipo da coluna"""
        if valor is None:
            return None

        if tipo_coluna in ['DATA', 'RESOLUCAO', 'NEGOCIACAO']:
            try:
                if isinstance(valor, (int, float)):
                    return pd.to_datetime('1899-12-30') + pd.Timedelta(days=int(valor))
                return pd.to_datetime(valor)
            except:
                return None
        
        if tipo_coluna == 'SITUACAO':
            try:
                return str(valor).strip().upper()
            except:
                return None

        return valor

def main():
    validador = ValidadorAvancado()
    
    arquivos = [
        Path("F:\\okok\\data\\(JULIO) LISTAS INDIVIDUAIS.xlsx"),
        Path("F:\\okok\\data\\(LEANDRO_ADRIANO) LISTAS INDIVIDUAIS.xlsx")
    ]
    
    for arquivo in arquivos:
        resultados = validador.processar_arquivo(arquivo)
        print(f"\nResultados para {arquivo.name}:")
        print(json.dumps(resultados, indent=2, ensure_ascii=False))

if __name__ == "__main__":
    main() 